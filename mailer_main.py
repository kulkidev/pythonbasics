import openpyxl as xl
import datetime
from utils.email.utils import send_email,build_message

html_email_template_part1 = """
<!DOCTYPE html>
    <html>
        <body>
            <p style="text-align: center"> Dear Manager, Hopefully this email finds you well.</p>
            <p style="text-align: center"> Please find below the list of employees reporting to you
            who have not logged in from last 90 days</p>
            <hr style="width: 500px;">
            <table>
                <thead>
                    <th>First Name</th>
                    <th>Last Name</th>
                    <th>Last Login</th>
                    <th>Email</th>
                </thead>
                <tbody>
"""
html_email_template_part2 = """
                </tbody>    
            </table>
        </body>
    </html>
"""


class Employee:
    def __init__(self,first_name, last_name, email, last_login_date, manager_email_id):
        self.first_name = first_name
        self.last_name = last_name
        self.full_name = first_name + " " + last_name
        self.email = email
        self.last_login_date = last_login_date
        self.manager_email_id = manager_email_id

    def __str__(self):
        return str(self.__class__) + '\n' + '\n'.join(
            (str(item) + ' = ' + str(self.__dict__[item]) for item in  sorted(self.__dict__)))


# This method reads data of employees from the excel spreadsheet
# and groups employees by their managers
def read_report(input_file_name, input_sheet_name):
    dormant_employee_dict = {}
    workbook = xl.load_workbook(input_file_name)
    sheet = workbook[input_sheet_name]

    print(sheet.max_row + 1)
    for row_number in range(2, sheet.max_row + 1):
        employee = Employee(sheet.cell(row=row_number, column=1).value,
                            sheet.cell(row=row_number, column=2).value,
                            sheet.cell(row=row_number, column=3).value,
                            sheet.cell(row=row_number, column=4).value,
                            sheet.cell(row=row_number, column=5).value)
        # print(f'employee : {employee.full_name} has manager {employee.manager_email_id}')
        if dormant_employee_dict.get(employee.manager_email_id):
            # print(f'manager : {employee.manager_email_id} exists')
            existing_employee_list = dormant_employee_dict.get(employee.manager_email_id)
            existing_employee_list.append(employee)
        else:
            # print(f'manager : {employee.manager_email_id} does not exist')
            new_list = [employee]
            dormant_employee_dict[employee.manager_email_id] = new_list

    return dormant_employee_dict


# This function connects to mail server and sends individual email
# to every manager using the credentials configured
def create_email_template_and_send(dormant_employees):
    for manager_email_id in dormant_employees:
        email_body_content = create_email_template(dormant_employees.get(manager_email_id))
        msg = build_message("pyemailtest111@gmail.com",
                   manager_email_id,
                   "De-activated user accounts summary",
                   email_body_content
                   )
        send_email(msg,
                   "Emailtest!1",
                   'smtp.gmail.com',
                   465)


# this function creates the HTML email body
# content of the email
def create_email_template(employee_list):
    html_content = ""
    html_str = ""
    for employee in employee_list:
        html_str = "<tr>" \
        + "<td>" + employee.first_name + "</td>"\
        + "<td>" + employee.last_name + "</td>"\
        + "<td>" + datetime.datetime.strftime(employee.last_login_date, '%Y-%m-%d') + "</td>"\
        + "<td>" + employee.email + "</td>"\
        + "</tr>"
        html_content += html_str

    return html_email_template_part1 + html_content + html_email_template_part2


create_email_template_and_send(read_report('data/expired_accounts.xlsx', 'Sheet1'))



